#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np

from tkinter import *
from tkinter.ttk import * 
from tkinter import messagebox

import os

import warnings
warnings.filterwarnings("ignore")

import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils  import get_column_letter
from openpyxl.styles.borders import Border, Side

import time
from tkinter import Radiobutton as rd
from tkinter import Entry as Entry_2
from tkinter import Scrollbar
import datetime as dt
#import xlwings as xw


# In[2]:


# wb = xw.Book("info.xlsx",password=9999)
# s = wb.sheets['Sheet1']

# info = s['A1:B5'].options(pd.DataFrame,index=False,header=False).value
# info

# wb.close()


# In[3]:


record = []
temp_record = []


# In[4]:


# total = 7
# criteria = 4

# h,m,s = "00","01","40"


# In[5]:


# info = pd.read_excel("info.xlsx",header=None)[1].values

# total = info[0]
# criteria = info[1]
# h = str(info[2])
# m = str(info[3])
# s = str(info[4])

# if len(h) <= 1:
#     h = "0" + h
    
# if len(m) <= 1:
#     m = "0" + m
    
# if len(s) <= 1:
#     s = "0" + s
    
# (h,m,s)


# In[6]:


id = ""

ans = ""

qn = ""

q_num = 0

var = ""

right = 0

wrong = 0

cat_val = ""
cat = ""

subcat_val = ""
sub = ""

name = ""

result = ""

temp = []

check = False

hour,minute,second = "","",""

total,criteria,h,m,s = "","","","",""
attemp = 0


st = ""

#admin
e_val = ""
del_emp = ""
del_st = ""
timer_st = ""
iv = ""
nv = ""
sv,sv_nick = "",""
qn_st = ""

tot_Question_val,pass_criteria_val,hour_val,min_val,sec_val = "","","","",""
del_qval,text_Answer,text_opt4,text_opt3,text_opt2,text_opt1,text_Question = "","","","","","",""


# In[7]:


def func():
    global window,q_num,qn,var,ans,right,wrong,name,cat,sub,temp,check,hour,minute,second,h,m,s,temp_record
     
    window.destroy()

    window = Tk()
    window.title('PTIS LMS')

    app_width = 950
    app_height = 600

    style = Style()
    
    
    hour=StringVar()
    minute=StringVar()
    second=StringVar()

    #setting the default value as 0
    hour.set(h)
    minute.set(m)
    second.set(s)
    
    
    style.configure('TButton', font =  ('calibri', 8))
    
    window.geometry(f'{app_width}x{app_height}+250+50')
#     window.overrideredirect(1)

    window.attributes('-toolwindow',True)
    window.protocol("WM_DELETE_WINDOW",submit)

#     window.attributes("-fullscreen",True)
    
    
    window.resizable(False,False)
    
    
    
    
    y = 5
    title = Label(window,text = "PTIS", font = ("Arial Bold",26),foreground = 'Red').place(x=400,y=y)
    y += 40
    
    sub_title = Label(window,text = f"Online Testing Module",font = ("Arial Bold",17)).place(x=330,y=y)

    y += 45

    ct = Label(window,text = f"Category: {cat}", font = ("Arial Bold",14),foreground = 'Green').place(x=20,y=y)
    
    y+=40
    
    hour_label = Label(window,textvariable = hour,font = ("calibri", 12),foreground="Red")


    mins_label = Label(window,textvariable = minute,font = ("calibri",12),foreground="Red")

    sec_label = Label(window,textvariable = second,font = ("calibri",12),foreground="Red")

    
    rem = Label(window,text = "Remaining Time",font = ("calibri",14),foreground="Blue")
    
    rem.place(x=49,y=12)


    rem = Label(window,text = f"Total Questions:   {str(total)}",font = ("calibri",13),foreground="Blue")
    
    rem.place(x=750,y=12)

    rem = Label(window,text = f"Attempted Qns:   {str(attemp)}",font = ("calibri",13),foreground="Blue")
    
    rem.place(x=750,y=34)

    rem = Label(window,text = f"Passing Criteria:  {str(criteria)}%",font = ("calibri",13),foreground="Blue")
    
    rem.place(x=750,y=56)
 
    
    hour_label.place(x=50,y=34)

    mins_label.place(x=100,y=34)

    sec_label.place(x=150,y=34)
    
    temp_record = []

    if len(str(qn[q_num][1])) > 400:
        text = str(q_num+1) +  ". " + str(qn[q_num][1])[:100] + "\n" + str(qn[q_num][1])[100:200] +"\n" +         str(qn[q_num][1])[200:300]+"\n" + str(qn[q_num][1])[300:400]+"\n" + str(qn[q_num][1])[400:]  
        if qn[q_num][1][100] == ' ':
                text = str(q_num+1)+  ". " + str(qn[q_num][1])[:100] + "\n" + str(qn[q_num][1])[100:]   
        else:
            number_he = 1
            while True:
                if qn[q_num][1][100 - number_he] == ' ':
                    text = str(q_num+1)+  ". " + str(qn[q_num][1])[:100 - number_he] + "\n" + str(qn[q_num][1])[100 - number_he:]
                    break
                number_he+=1

        if text[200] == ' ':
            text = str(text)[:200] + "\n" + str(text)[200:]   
        else:
            number_he = 1
            while True:
                if text[200 - number_he] == ' ':
                    text = str(text)[:200 - number_he] + "\n" + str(text)[200 - number_he:]
                    break
                number_he+=1

        if text[300] == ' ':
            text = str(text)[:300] + "\n" + str(text)[300:]   
        else:
            number_he = 1
            while True:
                if text[1][300 - number_he] == ' ':
                    text = str(text)[:300 - number_he] + "\n" + str(text)[300 - number_he:]
                    break
                number_he+=1

        if text[400] == ' ':
            text = str(text)[:400] + "\n" + str(text)[400:]   
        else:
            number_he = 1
            while True:
                if text[1][400 - number_he] == ' ':
                    text = str(text)[:400 - number_he] + "\n" + str(text)[400 - number_he:]
                    break
                number_he+=1

        
    elif len(str(qn[q_num][1])) > 300:
        if qn[q_num][1][100] == ' ':
                text = str(q_num+1)+  ". " + str(qn[q_num][1])[:100] + "\n" + str(qn[q_num][1])[100:]   
        else:
            number_he = 1
            while True:
                if qn[q_num][1][100 - number_he] == ' ':
                    text = str(q_num+1)+  ". " + str(qn[q_num][1])[:100 - number_he] + "\n" + str(qn[q_num][1])[100 - number_he:]
                    break
                number_he+=1

        if text[200] == ' ':
            text = str(text)[:200] + "\n" + str(text)[200:]   
        else:
            number_he = 1
            while True:
                if text[200 - number_he] == ' ':
                    text = str(text)[:200 - number_he] + "\n" + str(text)[200 - number_he:]
                    break
                number_he+=1

        if text[300] == ' ':
            text = str(text)[:300] + "\n" + str(text)[300:]   
        else:
            number_he = 1
            while True:
                if text[300 - number_he] == ' ':
                    text = str(text)[:300 - number_he] + "\n" + str(text)[300 - number_he:]
                    break
                number_he+=1

                
    elif len(str(qn[q_num][1])) > 200:

        if qn[q_num][1][100] == ' ':
                text = str(q_num+1)+  ". " + str(qn[q_num][1])[:100] + "\n" + str(qn[q_num][1])[100:]   
        else:
            number_he = 1
            while True:
                if qn[q_num][1][100 - number_he] == ' ':
                    text = str(q_num+1)+  ". " + str(qn[q_num][1])[:100 - number_he] + "\n" + str(qn[q_num][1])[100 - number_he:]
                    break
                number_he+=1

        if text[200] == ' ':
            text = str(text)[:200] + "\n" + str(text)[200:]   
        else:
            number_he = 1
            while True:
                if text[200 - number_he] == ' ':
                    text = str(text)[:200 - number_he] + "\n" + str(text)[200 - number_he:]
                    break
                number_he+=1
        
    
    elif len(str(qn[q_num][1])) > 100:
        c_no = 100
        if qn[q_num][1][c_no] == ' ':
                text = str(q_num+1)+  ". " + str(qn[q_num][1])[:c_no] + "\n" + str(qn[q_num][1])[c_no:]   
        else:
            number_he = 1
            while True:
                if qn[q_num][1][c_no - number_he] == ' ':
                    text = str(q_num+1)+  ". " + str(qn[q_num][1])[:c_no - number_he] + "\n" + str(qn[q_num][1])[c_no - number_he:]
                    break
                number_he+=1 
    
    else:
        text = str(q_num+1)+  ". " + str(qn[q_num][1])
    
    
#     text = str(q_num+1)+  ". " + str(qn[q_num][1])[:70] + "\n" + str(qn[q_num][1])[70:] 
#     text = str(q_num+1)+  ". " + str(qn[q_num][1])
    #q_label = Label(window,text = text,font = ("calibri",15))

    q_label = Text(window,width=88, height=3,font = ("calibri",14))
    q_label.configure(state='normal')
    q_label.insert('end', text)
    q_label.configure(state='disabled')

    scroll_q = Scrollbar(window, command=q_label.yview)
    q_label.configure(yscrollcommand=scroll_q.set)
    
    temp_record.append(text)

    var = StringVar()
    var.set(0)
    if len(str(qn[q_num][2])) > 400:
        text = str(qn[q_num][2])[:100] + "\n" + str(qn[q_num][2])[100:200] +"\n" +         str(qn[q_num][2])[200:300]+"\n" + str(qn[q_num][2])[300:400]+"\n" + str(qn[q_num][2])[400:]  

    elif len(str(qn[q_num][2])) > 300:
        text = str(qn[q_num][2])[:100] + "\n" + str(qn[q_num][2])[100:200] +"\n" +         str(qn[q_num][2])[200:300]+"\n" + str(qn[q_num][2])[300:]     
    
    elif len(str(qn[q_num][2])) > 200:
        text = str(qn[q_num][2])[:100] + "\n" + str(qn[q_num][2])[100:200] +"\n" + str(qn[q_num][2])[200:] 
    
    elif len(str(qn[q_num][2])) > 100:
        text = str(qn[q_num][2])[:100] + "\n" + str(qn[q_num][2])[100:]   
    
    else:
        text = str(qn[q_num][2])
    
    #text = str(qn[q_num][2])

    rb1 = rd(window,text=text,variable = var,value='A',command = answer,             font = ("calibri 14"),activeforeground="blue",borderwidth=5,justify="left",height =3 )

    temp_record.append(text)
    
    if len(str(qn[q_num][3])) > 400:
        text = str(qn[q_num][3])[:100] + "\n" + str(qn[q_num][3])[100:200] +"\n" +         str(qn[q_num][3])[200:300]+"\n" + str(qn[q_num][3])[300:400]+"\n" + str(qn[q_num][3])[400:]  

    elif len(str(qn[q_num][3])) > 300:
        text = str(qn[q_num][3])[:100] + "\n" + str(qn[q_num][3])[100:200] +"\n" +         str(qn[q_num][3])[200:300]+"\n" + str(qn[q_num][3])[300:]     
    
    elif len(str(qn[q_num][3])) > 200:
        text = str(qn[q_num][3])[:100] + "\n" + str(qn[q_num][3])[100:200] +"\n" + str(qn[q_num][3])[200:] 
    
    elif len(str(qn[q_num][3])) > 100:
        text = str(qn[q_num][3])[:100] + "\n" + str(qn[q_num][3])[100:]   
    
    else:
        text = str(qn[q_num][3])  

    #text = str(qn[q_num][3])  
    temp_record.append(text)
    
    rb2 = rd(window,text=text,variable = var,value='B',command = answer,             font = ("calibri 14"),activeforeground="blue",borderwidth=5,justify="left",height =3 )

    if len(str(qn[q_num][4])) > 400:
        text = str(qn[q_num][4])[:100] + "\n" + str(qn[q_num][4])[100:200] +"\n" +         str(qn[q_num][4])[200:300]+"\n" + str(qn[q_num][4])[300:400]+"\n" + str(qn[q_num][4])[400:]  

    elif len(str(qn[q_num][4])) > 300:
        text = str(qn[q_num][4])[:100] + "\n" + str(qn[q_num][4])[100:200] +"\n" +         str(qn[q_num][4])[200:300]+"\n" + str(qn[q_num][4])[300:]     
    
    elif len(str(qn[q_num][4])) > 200:
        text = str(qn[q_num][4])[:100] + "\n" + str(qn[q_num][4])[100:200] +"\n" + str(qn[q_num][4])[200:] 
    
    elif len(str(qn[q_num][4])) > 100:
        text = str(qn[q_num][4])[:100] + "\n" + str(qn[q_num][4])[100:]   
    
    else:
        text = str(qn[q_num][4]) 

    #text = str(qn[q_num][4])
    temp_record.append(text)
    
    rb3 = rd(window,text=text,variable = var,value='C',command = answer,             font = ("calibri 14"),activeforeground="blue",borderwidth=5,justify="left",height =3)


    if len(str(qn[q_num][5])) > 400:
        text = str(qn[q_num][5])[:100] + "\n" + str(qn[q_num][5])[100:200] +"\n" +         str(qn[q_num][5])[200:300]+"\n" + str(qn[q_num][5])[300:400]+"\n" + str(qn[q_num][5])[400:]  

    elif len(str(qn[q_num][5])) > 300:
        text = str(qn[q_num][5])[:100] + "\n" + str(qn[q_num][5])[100:200] +"\n" +         str(qn[q_num][5])[200:300]+"\n" + str(qn[q_num][5])[300:]     
    
    elif len(str(qn[q_num][5])) > 200:
        text = str(qn[q_num][5])[:100] + "\n" + str(qn[q_num][5])[100:200] +"\n" + str(qn[q_num][5])[200:] 
    
    elif len(str(qn[q_num][5])) > 100:
        text = str(qn[q_num][5])[:100] + "\n" + str(qn[q_num][5])[100:]   
    
    else:
        text = str(qn[q_num][5])

    #text = str(qn[q_num][5])
    temp_record.append(text)
    
    
    rb4 = rd(window,text=text,variable = var,value='D',command = answer,             font = ("calibri 14"),activeforeground="blue",borderwidth=5,justify="left",height =3)
    
    #id_val = Entry(window,width = 18,font=('Arial 12'))
    
    
    q_label.place(x = 20,y = y)
    scroll_q.place(x=905, y=y,height = 71)
    #id_val.place(x = 190,y = y)
    y+=75


    rb1.place(x=20, y=y)
    y+=73

    rb2.place(x=20, y=y)
    y+=73

    rb3.place(x=20, y=y)
    y+=73

    rb4.place(x=20, y=y)
    y+=110
    
    
#     bt = Button(window,text = "NEXT",style='W.TButton')
#     bt.place(x = 400,y = 540)
    
#     if (q_num == total - 1) or (check):

#         if len(temp) == 0:
#             bt = Button(window,text = "SUBMIT",style='W.TButton')
#             bt.place(x = 400,y = 540)
            
            
    style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')
    
    ans = ""
    t = int(hour.get())*3600 + int(minute.get())*60 + int(second.get())
    while t >= -1:
        
        mins,secs = divmod(t,60)
        hours=0
        if mins >60:
            
            hours, mins = divmod(mins, 60)
        
        
        hours = str(hours)
        mins = str(mins)
        secs = str(secs)
        
        if len(hours) <= 1:
            hour.set("{:0>2d}".format(int(hours)))
        else:
            hour.set("{}".format(int(hours)))
        
        if len(mins) <= 1:
            minute.set("{:0>2d}".format(int(mins)))
        else:
            minute.set("{}".format(int(mins)))
            
        if len(secs) <= 1:
            second.set("{:0>2d}".format(int(secs)))
        else:
            second.set("{}".format(int(secs)))
        
#         hour.set(hours)
#         minute.set(mins)
#         second.set(secs)
        
        if (t == 0):
            #messagebox.showinfo("Time Countdown", "Time's up ")
            ans = "-"
            submit()
            break
            
            
        if (q_num != total - 1) and (check == False):
            h = hour.get()
            m =  minute.get()
            s = second.get()
            s_bt = Button(window,command = skip,text = "SKIP",style='W.TButton')
            s_bt.place(x = 820,y = 540)
            

        if (q_num == total - 1) or (check):

            if len(temp) == 0:
                h = hour.get()
                m =  minute.get()
                s = second.get()

                bt = Button(window,command = submit,text = "SUBMIT",style='W.TButton')
                bt.place(x = 400,y = 540)
                

            elif len(temp) != 0:
                h = hour.get()
                m =  minute.get()
                s = second.get()
                bt = Button(window,command = next2,text = "NEXT",style='W.TButton')
                bt.place(x = 400,y = 540)
                


        else:

            bt = Button(window,command = next,text = "NEXT",style='W.TButton')
            bt.place(x = 400,y = 540)
            
        
        window.update()
        time.sleep(1)
        t -= 1
        
        
    #pos
    # file.place(x=20,y=100,rely = -.005)
    # file_val.place(x = 250,y = 100)
    

    window.mainloop()


# In[8]:


def skip():
    global temp,qn,q_num
    
    temp.append(q_num)
    q_num+=1
    
    
    return func()

def next2():
    global attemp,window,q_num,qn,var,ans,right,wrong,name,cat,sub,temp,check,record,temp_record
    
    check = True
    if ans != "":
        
        if qn[q_num][-2] == ans:
            right += 1
            
        else:
            wrong += 1
            
            
        temp_record.extend([qn[q_num][-2],ans])
        record.append(temp_record)
        attemp += 1
        
        q_num = temp.pop(0)
        return func()
        
    else:
        return messagebox.showinfo(title = "Error", message = 'Select any option')
            
    


# In[9]:


# def test_record():
#     global record
#     record_df = pd.DataFrame(record,columns = ["Question","Option A","Option B","Option C","Option D","Answer Key","User Answer"])
#     try:
#         wb = op.load_workbook(f"Answer Sheet\\{name}.xlsx")
#         no = len(wb.get_sheet_names()) + 1
#         ws = wb.create_sheet()
#         ws.title = f"Test {no}"

#     except:
#         wb = op.Workbook()
#         ws = wb.active

#         ws.title = "Test 1"

#     ws.sheet_view.showGridLines = False
#     ws['A1'] = f"Standard: {sub} {cat}"
#     ws['A1'].font = op.styles.Font(name = "Times New Roman",bold=True,size=11)    

#     ws['A3'] = f"Name: {name}"
#     ws['A3'].font = op.styles.Font(name = "Times New Roman",bold=True,size=11)

#     ws['A4'] = f"Total Questions:    {total}"
#     ws['A4'].font = op.styles.Font(name = "Times New Roman",bold=True,size=11)

#     ws['A5'] = f"Correct Answers:  {right}"
#     ws['A5'].font = op.styles.Font(name = "Times New Roman",bold=True,size=11)

#     ws['A6'] = f"Wrong Answers:    {wrong}"
#     ws['A6'].font = op.styles.Font(name = "Times New Roman",bold=True,size=11)

#     ws['A7'] = f"Status: {st}"
#     ws['A7'].font = op.styles.Font(name = "Times New Roman",bold=True,size=11)

    
#     rows = dataframe_to_rows(record_df,index=False)

#     for i, row in enumerate(rows,10):
#         for j,value in enumerate(row,1):


#             thin_border = Border(left = Side(style='thin'),
#                                  right = Side(style='thin'),
#                                  top = Side(style='thin'),
#                                  bottom=Side(style='thin'))

#             ws.cell(row=i,column=j,value=value).border = thin_border

#             dim = get_column_letter((j))

#             ws[f"{dim}{i}"].font = op.styles.Font(name = "Times New Roman",size=11)

#             if i == 10:
#                 ws[f"{dim}{i}"].font = op.styles.Font(name = "Times New Roman",bold=True,size=11)

#     wb.save(f"Answer Sheet\\{name}.xlsx")


# In[10]:


def answer():
    
    global ans,var
    ans = var.get()
    
    print(ans)

def disable():
    pass
    
def next():
    global window,q_num,qn,var,ans,right,wrong,name,cat,sub,record,temp_record,attemp
    if ans != "":
        
        if qn[q_num][-2] == ans:
            right += 1
            
        else:
            wrong += 1
            
        temp_record.extend([qn[q_num][-2],ans])
        record.append(temp_record)
        
        q_num += 1
        
        attemp += 1
        return func()
        
    else:
        return messagebox.showinfo(title = "Error", message = 'Select any option')
    
def submit_close():
    global window

    window.destroy()

def submit():
    global window,q_num,qn,var,ans,right,wrong,id,criteria,result,cat,sub,name,record,temp_record,st
    try:
        
#         if ans != "":

        if qn[q_num][-2] == ans:
            right += 1

        else:
            wrong += 1

        
        st = ""

        per = (right/total)*100
        if per >=  float(criteria):
            st = "Pass"
        else:
            st = "Fail"
        
        type = f"{sub} {cat}"
        #ind =  result[result["ID"] == id].index[0]
        #result.iloc[ind,:] = [id,name,total,right,wrong,st,type]
        s = time.ctime()
        s = dt.datetime.strptime(s,"%a %b  %d %H:%M:%S %Y").strftime("%d-%m-%Y %I:%M:%S %p")
        r = [int(id),name,int(total),int(right),int(wrong),str(round(per,2))+"%",str(criteria)+"%",st,type,s]


        wb = op.load_workbook(f"{result_db_loc}\\Result 2.xlsx")
        #ws = wb[f"{sub} {cat}"]
        db = pd.read_excel(f"{result_db_loc}\\Result 2.xlsx",sheet_name = "Result")
        #db = pd.concat(db,[r],ignore_index=True)
        db = pd.concat([db,pd.DataFrame([r],columns = db.columns)],ignore_index=True)
        
        ws = wb["Result"]
        rows = dataframe_to_rows(db,index=False)
        for i, row in enumerate(rows,1):
            for j,value in enumerate(row,1):
                thin_border = Border(left = Side(style='thin'),
                                     right = Side(style='thin'),
                                     top = Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.cell(row=i,column=j,value=value).border = thin_border
        wb.save(f"{result_db_loc}\\Result 2.xlsx")


        window.destroy()

    except:
        wrong -= 1
        return messagebox.showinfo(title = "Error", message = 'Something went wrong!!!')

    try:
            window = Tk()
            window.title('PTIS LMS')

            app_width = 550
            app_height = 420

            style = Style()
            style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')

            window.geometry(f'{app_width}x{app_height}+380+200')
            window.resizable(False,False)
            window.protocol("WM_DELETE_WINDOW",submit_close)

            y = 5
            title = Label(window,text = "PTIS", font = ("Arial Bold",26),foreground = 'Red').place(x=230,y=y)
            y += 40
            sub_title = Label(window,text = "TEST RESULT",font = ("Arial Bold",17)).place(x=190,y=y)
            y+=70

            id_l = Label(window,text = f"ID:",font = ("Arial",14),foreground = 'Blue').place(x=50,y=y)
            id_la = Label(window,text = id,font = ("Arial",14)).place(x=250,y=y)

            y+=40

            id_n = Label(window,text = f"Name:",font = ("Arial",14),foreground = 'Blue').place(x=50,y=y)
            id_na = Label(window,text = name,font = ("Arial",14)).place(x=250,y=y)

            y+=40

            id_q = Label(window,text = f"Total Question:",font = ("Arial",14),foreground = 'Blue').place(x=50,y=y)
            id_qa = Label(window,text = total,font = ("Arial",14)).place(x=250,y=y)

            y+=40

            id_r = Label(window,text = f"Correct Answer:",font = ("Arial",14),foreground = 'Blue').place(x=50,y=y)
            id_ra = Label(window,text = right,font = ("Arial",14)).place(x=250,y=y)

            y+=40

            id_w = Label(window,text = f"Percentage:",font = ("Arial",14),foreground = 'Blue').place(x=50,y=y)
            id_wa = Label(window,text = str(round(per,2))+"%",font = ("Arial",14)).place(x=250,y=y)

            y+=40

            id_cr = Label(window,text = f"Passing Criteria %:",font = ("Arial",14),foreground = 'Blue').place(x=50,y=y)
            id_cra = Label(window,text = str(criteria) + "%",font = ("Arial",14)).place(x=250,y=y)

            y+=40

            id_p = Label(window,text = f"Status:",font = ("Arial",14),foreground = 'Blue').place(x=50,y=y)
            if per >=  float(criteria):
                id_pa = Label(window,text = "Pass",font = ("Arial",14)).place(x=250,y=y)
            else:
                id_pa = Label(window,text = "Fail",font = ("Arial",14)).place(x=250,y=y)

            
            
            temp_record.extend([qn[q_num][-2],ans])
            record.append(temp_record)
        
        
        #test_record()
            window.mainloop()
#         else:
#             return messagebox.showinfo(title = "Error", message = 'Something went wrong!!!')
    except:
            pass
        
    #except:
     #   wrong -= 1
      #  return messagebox.showinfo(title = "Error", message = 'Something went wrong!!!')


# ### Admin

# In[11]:


def admin():
    global window,df,e_val,del_emp,del_st,timer_st,qn_st,df,standard
    

    window.destroy()
    
    df = pd.read_excel(f"{result_db_loc}\\Result 2.xlsx",sheet_name="Emloyees Data")
    standard = pd.read_excel(f"{result_db_loc}\\Result 2.xlsx",sheet_name="Standard")

    window = Tk()
    window.title('PTIS LMS')

    app_width = 510
    app_height = 420

    style = Style()
    style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')

    window.geometry(f'{app_width}x{app_height}+380+200')
    window.resizable(False,False)

    y = 10
    title = Label(window,text = "PTIS", font = ("Arial Bold",26),foreground = 'Red').place(x=200,y=y)
    y += 40
    sub_title = Label(window,text = "TEST",font = ("Arial Bold",17)).place(x=205,y=y)
    y+=30
    
    sub_title = Label(window,text = "Admin",font = ("Arial",15),foreground = 'Green').place(x=210,y=y)
    y+=50
    
    add_emp = Label(window,text = "Employee",font = ("calibri",15))
    del_emp = Combobox(window,width = 20,font=('Arial',11)) 
    a = list(df.iloc[:,1].values)
    del_emp['values'] = sorted(a)
    emp_del_bt = Button(window,command = employee_del,text = "Delete")
    emp_add_bt = Button(window,command = employee_add,text = "Add")
    
    add_st = Label(window,text = "Standard",font = ("calibri",15))
    del_st = Combobox(window,width = 20,font=('Arial',11)) 
    del_st['values'] = sorted(list(standard.iloc[:,0].values))
    st_del_bt = Button(window,text = "Delete",command = standard_del)
    st_add_bt = Button(window,text = "Add",command = standard_add)
    
    timer = Label(window,text = "Timer",font = ("calibri",15))
    timer_st = Combobox(window,width = 20,font=('Arial',11)) 
    timer_st['values'] = sorted(list(standard.iloc[:,0].values))
    timer_bt = Button(window,text = "Set",command = timer_add)
    
    add_q = Label(window,text = "Set Question",font = ("calibri",15))
    qn_st = Combobox(window,width = 20,font=('Arial',11)) 
    qn_st['values'] = sorted(list(standard.iloc[:,0].values))
    q_bt = Button(window,text = "Set",command=set_Question)
    dq_bt = Button(window,text = "Download Qns",command=download_question)
    
    
    e = Label(window,text = "Export Result",font = ("calibri",15))
    e_val = Combobox(window,width = 20,font=('Arial',11)) 
    e_val['values'] = sorted(a)


    bt = Button(window,command = export,text = "EXPORT")

    #pos
    # file.place(x=20,y=100,rely = -.005)
    # file_val.place(x = 250,y = 100)

    add_emp.place(x = 20,y = y)
    del_emp.place(x = 150,y = y+2)
    emp_del_bt.place(x = 340,y = y+1)
    emp_add_bt.place(x = 420,y = y+1)
    y+=50
    
    add_st.place(x = 20,y = y)
    del_st.place(x = 150,y = y+2)
    st_del_bt.place(x = 340,y = y+1)
    st_add_bt.place(x = 420,y = y+1)
    y+=50
    
    
    timer.place(x = 20,y = y)
    timer_st.place(x = 150,y = y+2)
    timer_bt.place(x = 340,y = y+1)
    y+=50
    
    add_q.place(x = 20,y = y)
    qn_st.place(x = 150,y = y+2)
    q_bt.place(x = 340,y = y+1)
    dq_bt.place(x = 420,y = y+1)
    y+=50
    
    e.place(x = 20,y = y)
    e_val.place(x = 150,y = y+2)
    bt.place(x = 340,y = y+1)
    #y+=50

    #bt.place(x = 170,y = 360)

    window.mainloop()


# In[12]:


def standard_add():
    global sv,sv_nick
    window_emp = Tk()
    window_emp.title('PTIS LMS')

    app_width = 460
    app_height = 250

    style = Style()
    style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')

    window_emp.geometry(f'{app_width}x{app_height}+380+200')
    window_emp.resizable(False,False)
    
    
    add_id = Label(window_emp,text = "Standard",font = ("calibri",15))
    sv = Entry(window_emp, font = ('Bold',12),width=28)

    add_nick = Label(window_emp,text = "Short Name",font = ("calibri",15))
    sv_nick = Entry(window_emp, font = ('Bold',12),width=28)
    
    def update1(event):
        leng_st.config(text=str(len(sv.get())))

    leng_st = Label(window_emp,text = "0" ,font = ("calibri",12))

    def update2(event):
        leng_sn.config(text=str(len(sv_nick.get())))

    leng_sn = Label(window_emp,text = "0" ,font = ("calibri",12))
    

    note = Label(window_emp,text = "Note:\nMake Short Name for your Standard if standard name length is >31.\nIf Standard length is <31 then copy paste standard in Short Name.\nShort Name length must be less than 31",font = ("calibri",11),foreground = 'Red')
    b = Button(window_emp,text = "ADD",command=add_standards)
    sv.bind('<KeyPress>', update1)
    sv.bind('<KeyRelease>', update1)

    sv_nick.bind('<KeyPress>', update2)
    sv_nick.bind('<KeyRelease>', update2)
    
    y = 40
    add_id.place(x=20,y=y)
    sv.place(x=150,y=y+2)
    leng_st.place(x=420,y=y+2)
    y += 40
    add_nick.place(x=20,y=y)
    sv_nick.place(x=150,y=y+2)
    leng_sn.place(x=420,y=y+2)
    
    y+=50
    b.place(x=190,y=y+2)
    y+=40
    note.place(x=10,y=y+2)
    

def add_standards():
    global standard
    
    df = pd.concat([standard,pd.DataFrame([[sv.get(),sv_nick.get()]],columns = standard.columns)],ignore_index=True)
    try:       
        wb = op.load_workbook(f"{result_db_loc}\\Result 2.xlsx")

        ws = wb["Standard"]

        rows = dataframe_to_rows(df,index=False)
        
        for i, row in enumerate(rows,1):
            for j,value in enumerate(row,1):


                thin_border = Border(left = Side(style='thin'),
                                     right = Side(style='thin'),
                                     top = Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.cell(row=i,column=j,value=value).border = thin_border

                dim = get_column_letter((j))

                
        ws[f"A1"].font = op.styles.Font(bold=True,size=11)
        ws[f"B1"].font = op.styles.Font(bold=True,size=11)

        wb.save(f"{result_db_loc}\\Result 2.xlsx")
        
        
        information = pd.DataFrame([["Total Questions",0],
                                    ["Passing Criteria",0],
                                    ["hours","00"],
                                    ["minutes","00"],
                                    ["seconds","00"]])
        
        wb = op.load_workbook(f"{info_db_loc}\\info.xlsx")
        ws = wb.create_sheet(sv_nick.get())
        rows = dataframe_to_rows(information,index=False,header=False)
        for i, row in enumerate(rows,1):
            for j,value in enumerate(row,1):
                thin_border = Border(left = Side(style='thin'),
                                     right = Side(style='thin'),
                                     top = Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.cell(row=i,column=j,value=value).border = thin_border
                
                dim = get_column_letter((j))

                ws[f"{dim}{i}"].font = op.styles.Font(size=18)
        
        wb.save(f"{info_db_loc}\\info.xlsx")
        
        messagebox.showinfo(title = "Done", message = "New Standard Added!")
    
        admin()
        
    except:
        messagebox.showinfo(title = "Error", message = "Something went wrong!")
    

    
    
def standard_del():
    global standard
    if del_st.get() not in standard.iloc[:,0].values:
        return messagebox.showinfo(title = "Error", message = "Standard Not Found In Database!")
                                  
    else:
        temp_sheet = standard[standard.iloc[:,0] == del_st.get()].iloc[:,1].values[0]
        standard = standard[standard.iloc[:,0] != del_st.get()]

        try:       
            wb = op.load_workbook(f"{result_db_loc}\\Result 2.xlsx")
            ws = wb["Standard"]
            ws.delete_cols(2)
            ws.delete_cols(1)
            rows = dataframe_to_rows(standard,index=False)

            for i, row in enumerate(rows,1):
                for j,value in enumerate(row,1):
                    thin_border = Border(left = Side(style='thin'),
                                         right = Side(style='thin'),
                                         top = Side(style='thin'),
                                         bottom=Side(style='thin'))

                    ws.cell(row=i,column=j,value=value).border = thin_border
                    dim = get_column_letter((j))

            ws["A1"].font = op.styles.Font(bold=True,size=11)
            ws["B1"].font = op.styles.Font(bold=True,size=11) 
            wb.save(f"{result_db_loc}\\Result 2.xlsx")

            wb = op.load_workbook(f"{info_db_loc}\\info.xlsx")
            del wb[temp_sheet]
            wb.save(f"{info_db_loc}\\info.xlsx")
            
            messagebox.showinfo(title = "Done", message = f"{del_st.get()} Deleted!")
            
            admin()
        except:
            messagebox.showinfo(title = "Error", message = "Something went wrong!")


# In[13]:


def add_nameid():
    global df
    
    print(iv.get(),nv.get())
    
    df = pd.concat([df,pd.DataFrame([[iv.get(),nv.get()]],columns = df.columns)],ignore_index=True)
    try:       
        wb = op.load_workbook(f"{result_db_loc}\\Result 2.xlsx")

        ws = wb["Emloyees Data"]

        rows = dataframe_to_rows(df,index=False)
        
        for i, row in enumerate(rows,1):
            for j,value in enumerate(row,1):


                thin_border = Border(left = Side(style='thin'),
                                     right = Side(style='thin'),
                                     top = Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.cell(row=i,column=j,value=value).border = thin_border

                dim = get_column_letter((j))

                
        ws[f"A1"].font = op.styles.Font(bold=True,size=11)
        ws[f"B1"].font = op.styles.Font(bold=True,size=11)

        wb.save(f"{result_db_loc}\\Result 2.xlsx")
        messagebox.showinfo(title = "Done", message = "New Employee Added!")
        
        admin()
        
    except:
        messagebox.showinfo(title = "Error", message = "Something went wrong!")
    


# In[14]:


def employee_add():
    global iv,nv
    window_emp = Tk()
    window_emp.title('PTIS LMS')

    app_width = 400
    app_height = 250

    style = Style()
    style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')

    window_emp.geometry(f'{app_width}x{app_height}+380+200')
    window_emp.resizable(False,False)

    
    add_id = Label(window_emp,text = "ID",font = ("calibri",15))
    iv = Entry(window_emp, font = ('Bold',12),width=25)
    
    add_name = Label(window_emp,text = "Name",font = ("calibri",15))
    nv = Entry(window_emp, font = ('Bold',12),width=25)
    
    b = Button(window_emp,text = "ADD",command=add_nameid)
    
    y = 20
    add_id.place(x=20,y=y)
    iv.place(x=100,y=y+2)
    
    y+=50
    add_name.place(x=20,y=y)
    nv.place(x=100,y=y+2)
    
    y+=50
    b.place(x=150,y=y+2)
    

def employee_del():
    global df
    if del_emp.get() not in df.iloc[:,1].values:
        return messagebox.showinfo(title = "Error", message = "Employee Not Found In Database!")
                                  
    else:                             
        df = df[df.iloc[:,1] != del_emp.get()]

        try:       
            wb = op.load_workbook(f"{result_db_loc}\\Result 2.xlsx")
            ws = wb["Emloyees Data"]
            rows = dataframe_to_rows(df,index=False)
            ws.delete_cols(1,2)
            for i, row in enumerate(rows,1):
                for j,value in enumerate(row,1):
                    thin_border = Border(left = Side(style='thin'),
                                         right = Side(style='thin'),
                                         top = Side(style='thin'),
                                         bottom=Side(style='thin'))

                    ws.cell(row=i,column=j,value=value).border = thin_border

            ws[f"A1"].font = op.styles.Font(bold=True,size=11)
            ws[f"B1"].font = op.styles.Font(bold=True,size=11)

            wb.save(f"{result_db_loc}\\Result 2.xlsx")
            messagebox.showinfo(title = "Done", message = f"{del_emp.get()} Deleted!")
            
            admin()
        except:
            messagebox.showinfo(title = "Error", message = "Something went wrong!")


# In[15]:


def export():
    global e_val
    print(e_val.get())
     
    d = pd.read_excel(f"{result_db_loc}\\Result 2.xlsx",sheet_name="Result")
    if len(d[d.iloc[:,1].astype(str) == e_val.get()]) > 0 :
        user = d[d.iloc[:,1].astype(str) == e_val.get()]
        
        user_n = d[d.iloc[:,1].astype(str) == e_val.get()].iloc[:,0].unique()[0]
        file_name = f"{user_n} - {e_val.get()}"
        #kuser.to_excel(f"Record\\{file_name}.xlsx",index=None)
        
        try:
            wb = op.load_workbook(f"{record_folder_loc}\\{file_name}.xlsx")
            #ws = wb.create_sheet()
        except:
            wb = op.Workbook()
            
        ws = wb.active    
        
        rows = dataframe_to_rows(user,index=False)
        for i, row in enumerate(rows,1):
            for j,value in enumerate(row,1):
                thin_border = Border(left = Side(style='thin'),
                                     right = Side(style='thin'),
                                     top = Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.cell(row=i,column=j,value=value).border = thin_border
                dim = get_column_letter((j))

                if i == 1:
                    ws.column_dimensions[dim].width = 20
                    ws[f"{dim}{i}"].font = op.styles.Font(bold=True,size=11)
    
        ws.auto_filter.ref = f"A1:J{str(len(user))}"
        wb.save(f"{record_folder_loc}\\{file_name}.xlsx")
        
        return messagebox.showinfo(title = "Done", message = "Done!!!")
    
    else:
        return messagebox.showinfo(title = "Error", message = "No Record")
        


# In[16]:


def timer_done():
    print(tot_Question_val.get(),pass_criteria_val.get(),hour_val.get(),min_val.get(),sec_val.get())
    
    d = pd.read_excel(f"{question_db_loc}\\Questions.xlsx")
    
    d = d[d.iloc[:,7]==timer_st.get()]
    
    if timer_st.get() != "Cummulative":    
        if len(d) < int(tot_Question_val.get()):
            return messagebox.showinfo(title = "Error", message = f"Questions in {timer_st.get()} is less than {tot_Question_val.get()}")
    
    if int(pass_criteria_val.get())/100 > int(tot_Question_val.get()):
        return messagebox.showinfo(title = "Error", message = "Passing Criteria should be less than Total Questions")
    
    try:
        temp_sheet = standard[standard.iloc[:,0]==timer_st.get()].iloc[:,1].values[0]
        
        wb = op.load_workbook(f"{info_db_loc}\\info.xlsx")
        ws = wb[temp_sheet]

        ws.delete_cols(2)

        ws["B1"] = int(tot_Question_val.get())
        ws["B1"].font = op.styles.Font(size=18)
        ws["B2"] = int(pass_criteria_val.get())
        ws["B2"].font = op.styles.Font(size=18)
        ws["B3"] = hour_val.get()
        ws["B3"].font = op.styles.Font(size=18)
        ws["B4"] = min_val.get()
        ws["B4"].font = op.styles.Font(size=18)
        ws["B5"] = sec_val.get()
        ws["B5"].font = op.styles.Font(size=18)
        
        wb.save(f"{info_db_loc}\\info.xlsx")
        return messagebox.showinfo(title = "Done", message = "Done")
        
    except:
        return messagebox.showinfo(title = "Error", message = "Something went Wrong")


def timer_add():
    global timer_st,tot_Question_val,pass_criteria_val,hour_val,min_val,sec_val
    if timer_st.get() != "":
        temp_sheet = standard[standard.iloc[:,0]==timer_st.get()].iloc[:,1].values[0]
        info_val = pd.read_excel(f"{info_db_loc}\\info.xlsx",sheet_name=temp_sheet,header=None)[1].values

        window_emp = Tk()
        window_emp.title('PTIS LMS')

        app_width = 400
        app_height = 380

        style = Style()
        style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')

        window_emp.geometry(f'{app_width}x{app_height}+380+200')
        window_emp.resizable(False,False)
    
        head = Label(window_emp,text = timer_st.get(),font = ("calibri",18,"bold"),foreground = 'Blue')

        tot_Question = Label(window_emp,text = "Total Question",font = ("calibri",15))
        tot_Question_val = Entry_2(window_emp, font = ('Bold',12),width=15)
        tot_Question_val.insert(END, str(info_val[0]))

        pass_criteria = Label(window_emp,text = "Passing Criteria %",font = ("calibri",15))
        pass_criteria_val = Entry_2(window_emp, font = ('Bold',12),width=15)
        pass_criteria_val.insert(END, str(info_val[1]))

        hour_label = Label(window_emp,text = "Set Hours",font = ("calibri",15))
        hour_val = Entry_2(window_emp, font = ('Bold',12),width=15)
        hour_val.insert(END, str(info_val[2]))

        min_label = Label(window_emp,text = "Set Minutes",font = ("calibri",15))
        min_val = Entry_2(window_emp, font = ('Bold',12),width=15)
        min_val.insert(END, str(info_val[3]))

        sec_label = Label(window_emp,text = "Set Seconds",font = ("calibri",15))
        sec_val = Entry_2(window_emp, font = ('Bold',12),width=15)
        sec_val.insert(END, str(info_val[4]))

        
        b = Button(window_emp,text = "Done",command = timer_done)

        y = 10
        
        head.place(x=50,y=y)
        y+=50
        
        tot_Question.place(x=20,y=y)
        tot_Question_val.place(x=180,y=y+2)

        y+=50
        pass_criteria.place(x=20,y=y)
        pass_criteria_val.place(x=180,y=y+2)

        y+=50
        hour_label.place(x=20,y=y)
        hour_val.place(x=180,y=y+2)

        y+=50
        min_label.place(x=20,y=y)
        min_val.place(x=180,y=y+2)

        y+=50
        sec_label.place(x=20,y=y)
        sec_val.place(x=180,y=y+2)

        y+=50
        b.place(x=150,y=y+2)


# In[17]:


def download_question():
    global qn_st
    print(qn_st.get())
     
    d = pd.read_excel(f"{question_db_loc}\\Questions.xlsx")
    d = d[d.iloc[:,7].astype(str) == qn_st.get()]
    if len(d) > 0:
        
        try:
            wb = op.load_workbook(f"{question_folder_loc}\\{qn_st.get()}.xlsx")
            #ws = wb.create_sheet()
        except:
            wb = op.Workbook()

        ws = wb.active    
        ws.delete_cols(1,8)
        rows = dataframe_to_rows(d,index=False)
        for i, row in enumerate(rows,1):
            for j,value in enumerate(row,1):
                thin_border = Border(left = Side(style='thin'),
                                     right = Side(style='thin'),
                                     top = Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.cell(row=i,column=j,value=value).border = thin_border
        wb.save(f"{question_folder_loc}\\{qn_st.get()}.xlsx")

        return messagebox.showinfo(title = "Done", message = "Done!!!")
    else:
        return messagebox.showinfo(title = "Error", message = f"No question in {qn_st.get()}!!!")
def set_Question():
    global del_qval,text_Answer,text_opt4,text_opt3,text_opt2,text_opt1,text_Question
    
    if qn_st.get() not in standard.iloc[:,0].values:
        return messagebox.showinfo(title = "Error", message = f"Select Correct Standard!!!")
    
    else:
        root = Tk()

        app_width = 550
        app_height = 600

        style = Style()
        style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')

        root.geometry(f'{app_width}x{app_height}+380+20')
        root.resizable(False,False)

        head = Label(root,text = "Add Questions",font = ("calibri",18,"bold"),foreground = 'Blue')
        y = 5

        head.place(x=210,y=y)
        y+=45
        #     text = Text(window, height=8, width=40)
        #     scroll = Scrollbar(window)
        #     text.configure(yscrollcommand=scroll.set)

        add_Question = Label(root,text = "Question",font = ("calibri",15))
        text_Question = Text(root, height=4, width=45, font = ('Bold',12))
        scroll_q = Scrollbar(root, command=text_Question.yview)
        text_Question.configure(yscrollcommand=scroll_q.set)

        add_Question.place(x=15,y=y+20)
        text_Question.place(x=100,y=y)
        scroll_q.place(x=510, y=y,height = 75)

        y+=90

        add_option1 = Label(root,text = "Option A",font = ("calibri",15))
        text_opt1 = Text(root, height=2, width=45, font = ('Bold',12))
        scroll_o1 = Scrollbar(root, command=text_opt1.yview)
        text_opt1.configure(yscrollcommand=scroll_o1.set)

        add_option1.place(x=15,y=y+7)
        text_opt1.place(x=100,y=y)
        scroll_o1.place(x=510, y=y,height = 40)

        y+=65

        add_option2 = Label(root,text = "Option B",font = ("calibri",15))
        text_opt2 = Text(root, height=2, width=45, font = ('Bold',12))
        scroll_o2 = Scrollbar(root, command=text_opt2.yview)
        text_opt2.configure(yscrollcommand=scroll_o2.set)

        add_option2.place(x=15,y=y+7)
        text_opt2.place(x=100,y=y)
        scroll_o2.place(x=510, y=y,height = 40)

        y+=65

        add_option3 = Label(root,text = "Option C",font = ("calibri",15))
        text_opt3 = Text(root, height=2, width=45, font = ('Bold',12))
        scroll_o3 = Scrollbar(root, command=text_opt3.yview)
        text_opt3.configure(yscrollcommand=scroll_o3.set)

        add_option3.place(x=15,y=y+7)
        text_opt3.place(x=100,y=y)
        scroll_o3.place(x=510, y=y,height = 40)

        y+=65

        add_option4 = Label(root,text = "Option D",font = ("calibri",15))
        text_opt4 = Text(root, height=2, width=45, font = ('Bold',12))
        scroll_o4 = Scrollbar(root, command=text_opt4.yview)
        text_opt4.configure(yscrollcommand=scroll_o4.set)

        add_option4.place(x=15,y=y+7)
        text_opt4.place(x=100,y=y)
        scroll_o4.place(x=510, y=y,height = 40)

        y+=65
        add_Answer = Label(root,text = "Answer",font = ("calibri",15))
        text_Answer = Combobox(root,width = 18,font=('Arial',11))                 
        text_Answer['values'] = ["A","B","C","D"]

        add_Answer.place(x=15,y=y)
        text_Answer.place(x=110,y=y)

        y+=40
        addq_bt = Button(root,text = "ADD",command=additon_question,width = 15)
        addq_bt.place(x=230,y=y)

        y+=60
        head = Label(root,text = "Delete Questions",font = ("calibri",18,"bold"),foreground = 'Blue')
        head.place(x=190,y=y)

        y+=50
        del_q = Label(root,text = "Question Number",font = ("calibri",15))
        del_qval = Entry_2(root, font = ('Bold',12),width=15)
        del_bt = Button(root,text = "Delete",width = 15,command=delete_question)

        del_q.place(x=50,y=y)
        del_qval.place(x=220,y=y+5)
        del_bt.place(x=400,y=y+4)
        root.mainloop()


# In[18]:


def additon_question():
    global text_Answer,text_opt4,text_opt3,text_opt2,text_opt1,text_Question
    
    d = pd.read_excel(f"{question_db_loc}\\Questions.xlsx")
    i = d.iloc[-1,0] + 1
    
    q = text_Question.get("1.0", "end-1c")
    q = q.replace("\n"," ")
    o1 = text_opt1.get("1.0", "end-1c")
    o1 = o1.replace("\n"," ")
    o2 = text_opt2.get("1.0", "end-1c")
    o2 = o2.replace("\n"," ")
    o3 = text_opt3.get("1.0", "end-1c")
    o3 = o3.replace("\n"," ")
    o4 = text_opt4.get("1.0", "end-1c")
    o4 = o4.replace("\n"," ")
    a = text_Answer.get()
    
    lst = [i,q,o1,o2,o3,o4,a,qn_st.get()]
    
    
    d = pd.concat([d,pd.DataFrame([lst],columns = d.columns)],ignore_index=True)
    
    
    try:       
        wb = op.load_workbook(f"{question_db_loc}\\Questions.xlsx")

        ws = wb.active

        rows = dataframe_to_rows(d,index=False)
        
        for i, row in enumerate(rows,1):
            for j,value in enumerate(row,1):


                thin_border = Border(left = Side(style='thin'),
                                     right = Side(style='thin'),
                                     top = Side(style='thin'),
                                     bottom=Side(style='thin'))

                ws.cell(row=i,column=j,value=value).border = thin_border

#                 dim = get_column_letter((j))

                
#         ws[f"A1"].font = op.styles.Font(bold=True,size=11)
#         ws[f"B1"].font = op.styles.Font(bold=True,size=11)

        wb.save(f"{question_db_loc}\\Questions.xlsx")
        messagebox.showinfo(title = "Done", message = "New Question Added!")
        
    except:
        messagebox.showinfo(title = "Error", message = "Something went wrong!")
    


# In[19]:


def delete_question():
    global del_qval,qn_st
    
    d = pd.read_excel(f"{question_db_loc}\\Questions.xlsx")
    
    a = d[d.iloc[:,7]==qn_st.get()]
    
    if "-" in del_qval.get():
            try:
                x = del_qval.get().split("-")
                r = list(range(int(x[0]),int(x[1])+1))
                d = d[d.iloc[:,0].astype(int).isin(r)!=True]

            except:
                return messagebox.showinfo(title = "Error", message = "Enter Correct input")
            
    elif del_qval.get() not in a.iloc[:,0].astype(int).astype(str).values :
        return messagebox.showinfo(title = "Error", message = f"Q.no: {del_qval.get()} not found !!!")
    
    else:
        d = d[d.iloc[:,0].astype(int).astype(str) != del_qval.get()]
    
    try:
         wb = op.load_workbook(f"{question_db_loc}\\Questions.xlsx")
            #ws = wb.create_sheet()
    except:
        wb = op.Workbook()

    ws = wb.active    
    ws.delete_cols(1,8)
    rows = dataframe_to_rows(d,index=False)
    for i, row in enumerate(rows,1):
        for j,value in enumerate(row,1):
            thin_border = Border(left = Side(style='thin'),
                                 right = Side(style='thin'),
                                 top = Side(style='thin'),
                                 bottom=Side(style='thin'))

            ws.cell(row=i,column=j,value=value).border = thin_border
    wb.save(f"{question_db_loc}\\Questions.xlsx")

    return messagebox.showinfo(title = "Done", message = "Done!!!")
    
#     except:
#         return messagebox.showinfo(title = "Error", message = "Something Went Wrong !!!")


# In[20]:


def intro():
    global id,window,q_num,qn,var,ans,cat_val,subcat_val,name,name_s,result,cat,sub,total,criteria,h,m,s
    
    try:
        name = name_s.get()
        id  = id_val.get()
        
        cat = cat_val.get()
        #sub = subcat_val.get()
        
        #category = sub + " " + cat
        category = cat
        print(category)
        a = pd.read_excel(f"{question_db_loc}\\Questions.xlsx")

        #if id == "9999" and cat_val.get() == "9999" and subcat_val.get() == "9999": 
        f = open(f"{password_loc}\\password.txt", "r")
        x = f.readline()
        f.close()

        
        if id == x and cat_val.get() == x:
            return admin()
            
                
        print(df.iloc[:,0].values)
        if id in df.iloc[:,0].astype(int).astype(str).values and (category in a["Standard"].values or category == "Cummulative"):
            info = pd.read_excel(f"{info_db_loc}\\info.xlsx",sheet_name=category,header=None)[1].values

            total = info[0]
            criteria = info[1]
            
            h = str(info[2])
            m = str(info[3])
            s = str(info[4])

            if len(h) <= 1:
                h = "0" + h

            if len(m) <= 1:
                m = "0" + m

            if len(s) <= 1:
                s = "0" + s
            
            if "Cummulative" in category:
                qn = a.copy()
            else:
                qn = a[a["Standard"] == category]
                
            qn.dropna(axis=1,how="all",inplace=True)   
            if (len(qn) < int(total)) or (len(qn) == 0) or (int(total) == 0):
                return messagebox.showinfo(title = "Error", message = f"Questions not defined for {category}")

            qn = qn.sample(total)
            qn = qn.values
            

            return func()
        
        else:
            return messagebox.showinfo(title = "Error", message = "Selction correct option.")

    except:
        
        return messagebox.showinfo(title = "Error", message = "Something Went Wrong") 
        
#     print(id)  
#     check = False
#     result["ID"] = result["ID"].astype("int").astype("str")
        
    


# In[21]:


def idname():
    global id_val,name_s,window,name_val,df
    
    
    if id_val.get() in df.iloc[:,0].astype("str").values:
        n = df.iloc[:,1][df.iloc[:,0].astype("str") == id_val.get()].values[0]
        name_s.set(n)
    else:
        n = ""
        name_s.set(n)
        messagebox.showinfo(title = "Error", message = 'INVALID ID')
        

def type_ok():
    global cat_val,window,standard

    cat_val['values'] = sorted(list(standard.iloc[:,0]))
    
#     print(name_s.get())
#     name_val.destroy()
#     name_val = Label(window,text = name_s.get(),font=('Arial 14')).place(x = 190,y = 180)


# In[22]:


#question_db_loc = "Z:\\PTIS Quiz\\DataBase"
# "\\\\110.93.228.131\\Sharing Folder\\Quiz\\\DataBase"

#question_db_loc =  "\\\\110.93.228.131\\Sharing Folder\\IT\\Ashar\\quiz db\\DataBase"
#question_db_loc = "E:\\Quiz new\\db"
question_db_loc = "E:\\Quiz Running\\db"
result_db_loc = question_db_loc
info_db_loc = question_db_loc
password_loc = question_db_loc


#question_folder_loc = "\\\\110.93.228.131\\Sharing Folder\\Quiz\\Admin\\Questions"
#question_folder_loc = "E:\\Quiz new"
question_folder_loc = "E:\\Quiz Running\\Questions"

#record_folder_loc = "\\\\110.93.228.131\\Sharing Folder\\Quiz\\Admin\\Record"
#record_folder_loc = "E:\\Quiz new"
record_folder_loc = "E:\\Quiz Running\\Record"

# In[ ]:


df = pd.read_excel(f"{question_db_loc}\\Result 2.xlsx",sheet_name="Emloyees Data")
standard = pd.read_excel(f"{question_db_loc}\\\Result 2.xlsx",sheet_name="Standard")
window = Tk()
window.title('PTIS LMS')

app_width = 450
app_height = 420

style = Style()
style.configure('W.TButton', font =  ('calibri', 15, 'bold'),foreground = 'Blue',background='Blue')

window.geometry(f'{app_width}x{app_height}+380+200')
window.resizable(False,False)

y = 10
title = Label(window,text = "PTIS", font = ("Arial Bold",26),foreground = 'Red').place(x=180,y=y)
y += 50
sub_title = Label(window,text = "Online Testing Module",font = ("Arial Bold",16)).place(x=120,y=y)
y+=100

name_s = StringVar()
name_s.set("")

id_label = Label(window,text = "ID",font = ("calibri",15))
id_val = Entry(window,width = 18,font=('Arial 12'))
id_bt = Button(window,command = idname,text = "OK")


name_label = Label(window,text = "Name",font = ("calibri",15))
name_val = Label(window,textvariable = name_s,font=('Arial 14'))


cat_label = Label(window,text = "Standard",font = ("calibri",15))
cat_val = Combobox(window,width = 18,font=('Arial',11))                 
cat_val['values'] = sorted(list(standard.iloc[:,0]))

# subcat_label = Label(window,text = "Category",font = ("calibri",15))
# subcat_val = Combobox(window,width = 18,font=('Arial',11))                 
# subcat_val['values'] = ["General","Specific"]

# level_label = Label(window,text = "Level",font = ("calibri",15))
# level_val =  Label(window,text = "II",font=('Arial 16'))


bt = Button(window,command = intro,text = "START TEST",style='W.TButton')

#pos
# file.place(x=20,y=100,rely = -.005)
# file_val.place(x = 250,y = 100)

id_label.place(x = 50,y = y)
id_val.place(x = 160,y = y)
#id_val.
id_bt.place(x = 345,y = y)
y+=40

name_label.place(x = 50,y = y)
name_val.place(x = 160,y = y)
y+=40


cat_label.place(x = 50,y = y)
cat_val.place(x = 160,y = y)
y+=60

# subcat_label.place(x = 50,y = y)
# subcat_val.place(x = 160,y = y)
# y+=40

# level_label.place(x = 50,y = y)
# level_val.place(x = 160,y = y)
# y+=60

bt.place(x = 170,y = 320)

window.mainloop()






